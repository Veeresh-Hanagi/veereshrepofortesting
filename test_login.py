import time
import unittest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import faker
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException

fake = faker.Faker()

file = "C:\\Users\\Veeresh\\Documents\\repo\\Orange_HRM_Login.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Login"]

# Generate fake usernames and passwords
for i in range(1, 15):
    sheet.cell(i, 1).value = fake.name()
    sheet.cell(i, 2).value = fake.password()

# Add valid login credentials to some rows
sheet.cell(4, 1).value = "Admin"
sheet.cell(4, 2).value = "admin123"
sheet.cell(6, 1).value = "Admin"
sheet.cell(6, 2).value = "admin123"
sheet.cell(8, 1).value = "Admin"
sheet.cell(8, 2).value = "admin123"

workbook.save(file)
rows = sheet.max_row

class loginfunctionality(unittest.TestCase):
    def setUp(self):
        self.driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        self.driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        self.driver.maximize_window()
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Username']"))
        )

    def test_loginfunctional(self):
        for i in range(1, rows + 1):
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Username']"))
                )
                username = self.driver.find_element(By.XPATH, "//input[@placeholder='Username']")
                password = self.driver.find_element(By.XPATH, "//input[@placeholder='Password']")
                login_button = self.driver.find_element(By.XPATH, "//*[@id='app']/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button")

                username.clear()
                password.clear()
                username.send_keys(sheet.cell(i, 1).value)
                password.send_keys(sheet.cell(i, 2).value)
                login_button.click()

                # Wait for alert OR dashboard user image
                try:
                    WebDriverWait(self.driver, 15).until(
                        EC.presence_of_element_located((By.XPATH, "//div[@role='alert']"))
                    )
                    alert_element = self.driver.find_element(By.XPATH, "//div[@role='alert']")
                    if alert_element.text.strip() == "Invalid credentials":
                        sheet.cell(i, 3).value = "Failed"
                        workbook.save(file)
                except TimeoutException:
                    # Assume login was successful
                    sheet.cell(i, 3).value = "Passed"
                    workbook.save(file)

                    # Perform logout
                    WebDriverWait(self.driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, "//img[@class='oxd-userdropdown-img']"))
                    ).click()

                    WebDriverWait(self.driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Logout']"))
                    ).click()

                    WebDriverWait(self.driver, 15).until(
                        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Username']"))
                    )

            except StaleElementReferenceException as e:
                print(f"StaleElementReferenceException at row {i}: {e}")
                sheet.cell(i, 3).value = "Error: Stale Element"
                workbook.save(file)
            except Exception as e:
                print(f"Unhandled exception at row {i}: {e}")
                sheet.cell(i, 3).value = f"Error: {str(e)}"
                workbook.save(file)

    def tearDown(self):
        self.driver.quit()
